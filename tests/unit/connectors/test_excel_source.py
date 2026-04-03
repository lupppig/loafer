"""Tests for ExcelSourceConnector."""

from __future__ import annotations

from typing import Any

import pytest

from loafer.connectors.sources.excel_source import ExcelSourceConnector
from loafer.exceptions import ExtractionError


def _create_xlsx(path: Any, rows: list[list[Any]], sheet_name: str = "Sheet1") -> None:
    """Helper: create a minimal .xlsx file."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    wb.save(str(path))
    wb.close()


class TestExcelSourceConnector:
    def test_basic_read(self, tmp_path: Any) -> None:
        f = tmp_path / "data.xlsx"
        _create_xlsx(f, [["id", "name"], [1, "Alice"], [2, "Bob"]])

        with ExcelSourceConnector(str(f)) as conn:
            rows = conn.read_all()

        assert len(rows) == 2
        assert rows[0] == {"id": 1, "name": "Alice"}

    def test_stream_chunked(self, tmp_path: Any) -> None:
        data = [["x"]] + [[i] for i in range(10)]
        f = tmp_path / "data.xlsx"
        _create_xlsx(f, data)

        with ExcelSourceConnector(str(f)) as conn:
            chunks = list(conn.stream(chunk_size=3))
            assert sum(len(c) for c in chunks) == 10

    def test_file_not_found(self, tmp_path: Any) -> None:
        with (
            pytest.raises(ExtractionError, match="not found"),
            ExcelSourceConnector(str(tmp_path / "nope.xlsx")) as _,
        ):
            pass

    def test_missing_sheet(self, tmp_path: Any) -> None:
        f = tmp_path / "data.xlsx"
        _create_xlsx(f, [["a"], [1]], sheet_name="Data")

        with (
            pytest.raises(ExtractionError, match="not found"),
            ExcelSourceConnector(str(f), sheet="Missing") as _,
        ):
            pass

    def test_empty_workbook(self, tmp_path: Any) -> None:
        import openpyxl

        f = tmp_path / "empty.xlsx"
        wb = openpyxl.Workbook()
        wb.save(str(f))
        wb.close()

        with ExcelSourceConnector(str(f)) as conn:
            rows = conn.read_all()

        assert rows == []

    def test_count(self, tmp_path: Any) -> None:
        f = tmp_path / "data.xlsx"
        _create_xlsx(f, [["x"], [1], [2], [3]])

        with ExcelSourceConnector(str(f)) as conn:
            assert conn.count() == 3

    def test_merged_cells(self, tmp_path: Any) -> None:
        import openpyxl

        f = tmp_path / "merged.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["region", "value"])
        ws.append(["North", 10])
        ws.append(["", 20])  # will be merged with above
        ws.merge_cells("A2:A3")
        wb.save(str(f))
        wb.close()

        with ExcelSourceConnector(str(f)) as conn:
            rows = conn.read_all()

        assert len(rows) == 2
        # Both rows should have the merged value "North".
        assert rows[0]["region"] == "North"
        assert rows[1]["region"] == "North"

    def test_named_sheet(self, tmp_path: Any) -> None:
        f = tmp_path / "data.xlsx"
        _create_xlsx(f, [["a"], [1]], sheet_name="MySheet")

        with ExcelSourceConnector(str(f), sheet="MySheet") as conn:
            rows = conn.read_all()

        assert len(rows) == 1
