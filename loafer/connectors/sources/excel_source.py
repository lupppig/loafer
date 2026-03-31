"""Excel source connector using openpyxl.

Reads computed values (not formulas), handles merged cells, validates
sheet names, and streams rows in chunks without buffering the full dataset.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import TYPE_CHECKING, Any

if TYPE_CHECKING:
    from collections.abc import Iterator

from loafer.connectors.base import SourceConnector
from loafer.exceptions import ExtractionError

logger = logging.getLogger(__name__)


class ExcelSourceConnector(SourceConnector):
    """Read rows from an Excel (.xlsx) file."""

    def __init__(self, path: str, sheet: str | None = None) -> None:
        self._path = Path(path)
        self._sheet = sheet
        self._wb: Any = None
        self._ws: Any = None
        self._headers: list[str] = []
        self._row_count: int | None = None

    def _load_workbook(self) -> Any:
        try:
            import openpyxl
        except ImportError as exc:  # pragma: no cover
            raise ExtractionError("openpyxl is required for Excel files") from exc
        try:
            return openpyxl.load_workbook(str(self._path), data_only=True)
        except Exception as exc:
            raise ExtractionError(f"Failed to open Excel file {self._path}: {exc}") from exc

    def connect(self) -> None:
        if not self._path.exists():
            raise ExtractionError(f"Excel file not found: {self._path}")

        self._wb = self._load_workbook()

        if self._sheet:
            if self._sheet not in self._wb.sheetnames:
                available = list(self._wb.sheetnames)
                self._wb.close()
                raise ExtractionError(f"Sheet '{self._sheet}' not found. Available: {available}")
            self._ws = self._wb[self._sheet]
        else:
            self._ws = self._wb.active
            if self._ws is None:  # pragma: no cover
                self._wb.close()
                raise ExtractionError("Workbook has no active sheet")

        rows_iter = self._ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            logger.warning("Excel file is empty: %s", self._path)
            self._headers = []
            return

        self._headers = [
            str(h) if h is not None else f"column_{i}" for i, h in enumerate(header_row)
        ]

        data_rows = self._ws.max_row - 1
        self._row_count = max(0, data_rows)

    def disconnect(self) -> None:
        if self._wb is not None:
            self._wb.close()
            self._wb = None
        self._ws = None
        self._headers = []

    def stream(self, chunk_size: int) -> Iterator[list[dict[str, Any]]]:
        if self._ws is None:
            raise ExtractionError("connect() must be called before stream()")

        if self._headers == []:
            return

        for merge_range in list(self._ws.merged_cells.ranges):
            min_row = merge_range.min_row
            min_col = merge_range.min_col
            top_left_value = self._ws.cell(row=min_row, column=min_col).value
            self._ws.unmerge_cells(str(merge_range))
            for row_idx in range(merge_range.min_row, merge_range.max_row + 1):
                for col_idx in range(merge_range.min_col, merge_range.max_col + 1):
                    self._ws.cell(row=row_idx, column=col_idx).value = top_left_value

        chunk: list[dict[str, Any]] = []
        total_rows = 0

        for row in self._ws.iter_rows(min_row=2, values_only=True):
            row_dict: dict[str, Any] = {}
            for header, value in zip(self._headers, row, strict=False):
                row_dict[header] = self._coerce_value(value)
            chunk.append(row_dict)
            total_rows += 1

            if len(chunk) >= chunk_size:
                yield chunk
                chunk = []

        if chunk:
            yield chunk

        self._row_count = total_rows

    def _coerce_value(self, value: Any) -> Any:
        if value is None:
            return None
        if isinstance(value, (int, float, str, bool)):
            return value
        return str(value)

    def count(self) -> int | None:
        return self._row_count
